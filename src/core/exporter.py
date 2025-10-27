"""
Excel exporter module for CIA Factbook JSON to Excel Exporter.

Handles conversion of parsed data to formatted Excel files using pandas and openpyxl.
"""

import logging
import os
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.config.config_loader import OUTPUT_DIR, OUTPUT_FILENAME, SHEET_NAME

# Configure logging
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Handles exporting parsed data to Excel format with formatting."""
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.output_dir = Path(OUTPUT_DIR)
        self.output_filename = OUTPUT_FILENAME
        self.sheet_name = SHEET_NAME
        
        # Ensure output directory exists
        self._ensure_output_directory()
    
    def _ensure_output_directory(self):
        """Create output directory if it doesn't exist."""
        try:
            self.output_dir.mkdir(exist_ok=True)
            logger.info(f"Output directory: {self.output_dir.absolute()}")
        except Exception as e:
            logger.error(f"Failed to create output directory: {str(e)}")
            raise
    
    def _get_output_path(self) -> Path:
        """Get the full path for the output Excel file."""
        return self.output_dir / self.output_filename
    
    def _create_dataframe(self, parsed_data: Dict[str, Dict[str, Optional[str]]]) -> pd.DataFrame:
        """
        Convert parsed data to pandas DataFrame.
        
        Args:
            parsed_data: Dictionary mapping country codes to field data
            
        Returns:
            pandas DataFrame with all country data
        """
        if not parsed_data:
            logger.warning("No data to export")
            return pd.DataFrame()
        
        # Convert to list of dictionaries for DataFrame creation
        records = list(parsed_data.values())
        
        # Create DataFrame
        df = pd.DataFrame(records)
        
        logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
        return df
    
    def _format_excel_file(self, file_path: Path):
        """
        Apply formatting to the Excel file.
        
        Args:
            file_path: Path to the Excel file to format
        """
        try:
            # Load the workbook
            workbook = load_workbook(file_path)
            worksheet = workbook[self.sheet_name]
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Format headers
            for column in worksheet.iter_cols(min_row=1, max_row=1):
                for cell in column:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Find the maximum length in the column
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze the header row
            worksheet.freeze_panes = "A2"
            
            # Save the formatted workbook
            workbook.save(file_path)
            logger.info("Applied Excel formatting successfully")
            
        except Exception as e:
            logger.error(f"Error applying Excel formatting: {str(e)}")
            # Don't raise here - formatting is optional
    
    def export_to_excel(self, parsed_data: Dict[str, Dict[str, Optional[str]]]) -> Optional[str]:
        """
        Export parsed data to Excel file with formatting.
        
        Args:
            parsed_data: Dictionary mapping country codes to field data
            
        Returns:
            Path to the created Excel file, or None if export failed
        """
        if not parsed_data:
            logger.error("No data to export")
            return None
        
        try:
            # Create DataFrame
            df = self._create_dataframe(parsed_data)
            
            if df.empty:
                logger.error("DataFrame is empty, nothing to export")
                return None
            
            # Get output path
            output_path = self._get_output_path()
            
            # Export to Excel
            logger.info(f"Exporting data to {output_path}")
            
            # Initial export with pandas
            with pd.ExcelWriter(
                output_path, 
                engine='openpyxl',
                mode='w'
            ) as writer:
                df.to_excel(
                    writer, 
                    sheet_name=self.sheet_name, 
                    index=False,
                    freeze_panes=(1, 0)  # Freeze header row
                )
            
            # Apply additional formatting
            self._format_excel_file(output_path)
            
            logger.info(f"Successfully exported {len(df)} countries to Excel")
            return str(output_path.absolute())
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            return None
    
    def get_export_summary(self, parsed_data: Dict[str, Dict[str, Optional[str]]]) -> Dict[str, any]:
        """
        Get summary information about the data to be exported.
        
        Args:
            parsed_data: Parsed country data
            
        Returns:
            Dictionary with export summary information
        """
        if not parsed_data:
            return {
                'total_countries': 0,
                'total_fields': 0,
                'field_names': [],
                'output_file': str(self._get_output_path().absolute())
            }
        
        # Get field names from first country
        first_country = next(iter(parsed_data.values()))
        field_names = list(first_country.keys())
        
        return {
            'total_countries': len(parsed_data),
            'total_fields': len(field_names),
            'field_names': field_names,
            'output_file': str(self._get_output_path().absolute())
        }
    
    def file_exists(self) -> bool:
        """Check if the output file already exists."""
        return self._get_output_path().exists()
    
    def get_file_size(self) -> Optional[int]:
        """Get the size of the output file in bytes."""
        file_path = self._get_output_path()
        if file_path.exists():
            return file_path.stat().st_size
        return None
